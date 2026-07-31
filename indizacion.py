"""
Determina en qué bases está indizada cada revista y deriva su nivel CONICET.

Fundamento — Resolución D 2249/2014 del Directorio del CONICET, anexo
"Bases para la Categorización de publicaciones periódicas para las Ciencias
Sociales y Humanidades según sus sistemas de indización":

  Nivel 1  Internacional Global y Regional: (Web of Science) ISI, SCOPUS,
           ERIH y demás índices nacionales de excelencia, SciELO.org, CIRC A
  Nivel 2  Internacional y Regional: Sage, Springer, Taylor & Francis, WILEY,
           Francis, Pascal, JSTOR, REDALyC, y Núcleo Básico de Revistas
           Científicas Argentinas
  Nivel 3  Bases Referenciales y Catálogos Regionales: Philosopher's Index,
           Sociological Abs., MLA, ERIC, PsycInfo, CLACSO, Latindex Catálogo…

Como todas las revistas de este catálogo pertenecen al Núcleo Básico, el piso
de todas es Nivel 2. Suben a Nivel 1 las que además estén en Scopus o SciELO.

ADVERTENCIA de la propia resolución: la jerarquización es de los *índices*, no
de cada revista. "Dentro de un mismo nivel o grupo conviven revistas que, si
bien de un nivel semejante en comparación con los otros, difieren entre sí
respecto de su calidad." No es un puntaje de calidad por revista.

Fuentes usadas:
  - Scopus Source List (Elsevier, xlsx oficial, actualizado mensualmente)
  - SciELO Argentina, nómina alfabética oficial de la colección
  - OpenAlex API (DOAJ, h-index, cantidad de trabajos)

No se consulta SCImago: su descarga está detrás de Cloudflare y no se evade.
El dato de Scopus —que es el que define el Nivel 1— sale de la lista oficial
de Elsevier, que es la misma base sobre la que SCImago calcula el SJR.

No se usa el campo `is_in_scielo` de OpenAlex: devuelve False para revistas
que están efectivamente alojadas en scielo.org.ar (verificado con Anclajes,
Avá y el Boletín Ravignani). Se usa la nómina de SciELO Argentina.
"""
import truststore
truststore.inject_into_ssl()

import os
import re
import time
import logging
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed

from database import (init_db, obtener_revistas, guardar_indizacion,
                      registrar_actualizacion)

logger = logging.getLogger(__name__)

DIR_DATOS = os.path.join(os.path.dirname(__file__), 'datos')
SCOPUS_XLSX = os.path.join(DIR_DATOS, 'scopus_source_list.xlsx')
SCOPUS_URL = ("https://downloads.ctfassets.net/o78em1y1w4i4/7xtaTxNiNcWRTeZkV86eNy/"
              "710bfd3c7f7c7c9c88eeb3638ba4be43/ext_list_Jun_2026.xlsx")

from configuracion import HEADERS  # el correo de contacto no va en el código

OPENALEX = "https://api.openalex.org/sources/issn:{}"
SCIELO_AR = ("https://www.scielo.org.ar/scielo.php?"
             "script=sci_alphabetic&lng=es&nrm=iso")


def normalizar_issn(issn):
    """'2250-4982' -> '22504982'. Devuelve '' si no es un ISSN."""
    if not issn:
        return ''
    s = re.sub(r'[^0-9Xx]', '', str(issn)).upper()
    return s if len(s) == 8 else ''


def descargar_scopus(forzar=False):
    """Baja la lista oficial de fuentes de Scopus si no está en disco."""
    os.makedirs(DIR_DATOS, exist_ok=True)
    if os.path.exists(SCOPUS_XLSX) and not forzar:
        return SCOPUS_XLSX
    logger.info("Descargando Scopus Source List (~26 MB)...")
    r = requests.get(SCOPUS_URL, headers={'User-Agent': 'Mozilla/5.0'}, timeout=300)
    r.raise_for_status()
    with open(SCOPUS_XLSX, 'wb') as f:
        f.write(r.content)
    return SCOPUS_XLSX


def cargar_scopus():
    """
    Devuelve {issn_normalizado: estado} de todas las fuentes Scopus.

    Se leen la hoja principal y la de títulos discontinuados: una revista que
    Scopus dejó de indizar no aparece en la principal, y omitirla haría parecer
    que nunca estuvo (le pasaba a Alteridades, Frontera Norte y Desacatos).
    """
    import openpyxl
    ruta = descargar_scopus()
    wb = openpyxl.load_workbook(ruta, read_only=True)

    mapa = {}
    for hoja in wb.sheetnames:
        h = hoja.lower()
        if not (h.startswith('scopus sources') or h.startswith('discontinued')):
            continue
        discontinuada = h.startswith('discontinued')
        ws = wb[hoja]
        filas = ws.iter_rows(values_only=True)
        try:
            cab = [str(c or '').strip().lower() for c in next(filas)]
        except StopIteration:
            continue
        if 'issn' not in cab:
            continue
        i_issn = cab.index('issn')
        i_eissn = cab.index('eissn') if 'eissn' in cab else None
        i_estado = next((i for i, c in enumerate(cab)
                         if c.startswith('active or inactive')), None)

        for f in filas:
            if i_estado is not None:
                estado = str(f[i_estado] or '').strip() or 'desconocido'
            else:
                estado = 'Discontinued'
            if discontinuada:
                estado = 'Discontinued'
            for idx in (i_issn, i_eissn):
                if idx is None:
                    continue
                n = normalizar_issn(f[idx])
                if n and mapa.get(n) != 'Active':
                    mapa[n] = estado
    wb.close()
    logger.info("Scopus: %d ISSN cargados (incluye discontinuados)", len(mapa))
    return mapa


def cargar_scielo():
    """
    ISSN de la colección SciELO Argentina, desde su nómina alfabética.
    Los enlaces del listado llevan el ISSN en el parámetro `pid`.
    """
    from bs4 import BeautifulSoup
    r = requests.get(SCIELO_AR, headers={'User-Agent': 'Mozilla/5.0'}, timeout=60)
    r.raise_for_status()
    soup = BeautifulSoup(r.content, 'html.parser')

    issns = set()
    for a in soup.find_all('a', href=re.compile(r'pid=\d{4}-\d{3}[\dXx]', re.I)):
        m = re.search(r'pid=(\d{4}-\d{3}[\dXx])', a['href'], re.I)
        if m:
            n = normalizar_issn(m.group(1))
            if n:
                issns.add(n)
    logger.info("SciELO Argentina: %d ISSN cargados", len(issns))
    return issns


def consultar_openalex(sesion, issns):
    """Consulta OpenAlex por los ISSN de una revista. Devuelve dict o {}.

    Solo se toman DOAJ y métricas: el campo is_in_scielo de OpenAlex es
    incorrecto para las revistas argentinas (ver docstring del módulo).
    """
    for issn in issns:
        if not issn:
            continue
        try:
            r = sesion.get(OPENALEX.format(issn), timeout=25)
            if r.status_code != 200:
                continue
            j = r.json()
            st = j.get('summary_stats') or {}
            return dict(en_doaj=1 if j.get('is_in_doaj') else 0,
                        openalex_core=1 if j.get('is_core') else 0,
                        h_index=st.get('h_index'),
                        works_count=j.get('works_count'))
        except (requests.RequestException, ValueError):
            continue
    return {}


def nivel_conicet(en_scopus, en_scielo):
    """
    Nivel según Res. 2249/14. Todas las revistas del catálogo están en el
    Núcleo Básico, que la resolución ubica en Nivel 2; Scopus y SciELO.org
    están nombrados en Nivel 1.
    """
    if en_scopus or en_scielo:
        return 1
    return 2


def actualizar_indizacion(workers=6, progreso=None):
    init_db()
    scopus = cargar_scopus()
    scielo = cargar_scielo()
    revistas = obtener_revistas()

    sesion = requests.Session()
    sesion.headers.update(HEADERS)

    resumen = dict(revisadas=0, nivel1=0, nivel2=0, scopus=0, scielo=0,
                   doaj=0, sin_openalex=0)

    def procesar(rv):
        issns_norm = [n for n in (normalizar_issn(rv['issn_impreso']),
                                  normalizar_issn(rv['issn_online'])) if n]
        estado_scopus = next((scopus[n] for n in issns_norm if n in scopus), None)
        en_scopus = 1 if estado_scopus else 0

        # SciELO: por ISSN en la nómina, o porque la revista está alojada allí.
        en_scielo = 1 if any(n in scielo for n in issns_norm) else 0
        if not en_scielo and rv['sitio_url'] and 'scielo.org' in rv['sitio_url'].lower():
            en_scielo = 1

        oa = consultar_openalex(
            sesion, [rv['issn_impreso'], rv['issn_online']])
        time.sleep(0.05)  # cortesía con la API pública de OpenAlex

        datos = dict(en_scopus=en_scopus, scopus_estado=estado_scopus,
                     en_scielo=en_scielo, en_doaj=oa.get('en_doaj'),
                     openalex_core=oa.get('openalex_core'),
                     h_index=oa.get('h_index'), works_count=oa.get('works_count'))
        datos['nivel_conicet'] = nivel_conicet(en_scopus, en_scielo)
        return rv, datos, bool(oa)

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futuros = [pool.submit(procesar, rv) for rv in revistas]
        for i, fut in enumerate(as_completed(futuros), 1):
            rv, datos, tuvo_oa = fut.result()
            guardar_indizacion(rv['id'], datos)

            resumen['revisadas'] += 1
            resumen['nivel1' if datos['nivel_conicet'] == 1 else 'nivel2'] += 1
            resumen['scopus'] += bool(datos['en_scopus'])
            resumen['scielo'] += bool(datos['en_scielo'])
            resumen['doaj'] += bool(datos['en_doaj'])
            if not tuvo_oa:
                resumen['sin_openalex'] += 1

            if progreso and i % 10 == 0:
                progreso(i, len(revistas))

    registrar_actualizacion(
        "indizacion",
        f"{resumen['revisadas']} revistas: {resumen['nivel1']} Nivel 1, "
        f"{resumen['nivel2']} Nivel 2 | Scopus {resumen['scopus']}, "
        f"SciELO {resumen['scielo']}, DOAJ {resumen['doaj']}")
    return resumen


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')
    r = actualizar_indizacion(progreso=lambda i, t: print(f"  {i}/{t}", flush=True))
    print("\nRESUMEN:")
    for k, v in r.items():
        print(f"  {k}: {v}")
