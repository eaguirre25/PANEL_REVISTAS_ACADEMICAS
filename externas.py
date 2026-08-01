"""
Incorpora revistas latinoamericanas de fuera del Núcleo Básico argentino.

Para cada revista del listado curado (datos_externas.py):
  1. Se busca el título en la lista oficial de Scopus (que trae título, ISSN y
     estado activo/discontinuado), y en DOAJ, que resuelve título -> ISSN.
  2. Con el ISSN obtenido se verifica la red SciELO completa (API ArticleMeta,
     2269 revistas de todas las colecciones).
  3. Se guarda tanto la indización *verificada* como la *declarada* en el
     listado de origen, para poder compararlas.

Las URLs que en el listado original eran búsquedas de Google NO se inventan:
se toman de DOAJ si están, o quedan vacías.

No se usa OpenAlex acá: su cuota gratuita diaria se agota con las ~324
consultas del catálogo NBRA y devuelve HTTP 429 el resto del día. DOAJ y la
lista de Scopus no tienen ese límite.

Sobre el nivel CONICET en revistas extranjeras: la Res. 2249/14 jerarquiza
bases de indización, así que Scopus y SciELO siguen implicando Nivel 1. Pero
el piso de Nivel 2 que tienen las revistas argentinas viene de pertenecer al
Núcleo Básico, y estas no pertenecen. Por eso, si no están en Scopus ni SciELO,
el nivel queda **sin determinar** en vez de asignarle uno inventado: haría
falta verificar REDALyC/JSTOR (Nivel 2) o Latindex Catálogo (Nivel 3), que no
publican listados abiertos comparables.
"""
import truststore
truststore.inject_into_ssl()

import re
import time
import logging
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed

from database import (init_db, guardar_revista_externa, guardar_indizacion,
                      registrar_actualizacion)
from indizacion import cargar_scopus, normalizar_issn
from datos_externas import REVISTAS

logger = logging.getLogger(__name__)

from configuracion import HEADERS  # el correo de contacto no va en el código

ARTICLEMETA = "https://articlemeta.scielo.org/api/v1/journal/identifiers/"
DOAJ_BUSCAR = "https://doaj.org/api/search/journals/title:{}"

PAISES = {
    "Chile": "CL", "Brasil": "BR", "Brasil/Portugal": "BR", "Uruguay": "UY",
    "Perú": "PE", "Paraguay": "PY", "Colombia": "CO", "Ecuador": "EC",
    "Venezuela": "VE", "Cuba": "CU", "Costa Rica": "CR", "Panamá": "PA",
    "El Salvador": "SV", "Guatemala": "GT", "Honduras": "HN", "Nicaragua": "NI",
    "México": "MX", "Bolivia": "BO", "Puerto Rico": "PR",
    "República Dominicana": "DO", "Argentina/regional": "AR",
    "Iberoamérica (OEI)": None,
}


def cargar_scielo_red():
    """ISSN de toda la red SciELO (todas las colecciones), vía ArticleMeta."""
    issns, offset = set(), 0
    while True:
        r = requests.get(ARTICLEMETA, headers=HEADERS, timeout=90,
                         params={'limit': 1000, 'offset': offset})
        r.raise_for_status()
        j = r.json()
        objetos = j.get('objects', [])
        for o in objetos:
            n = normalizar_issn(o.get('code'))
            if n:
                issns.add(n)
        total = (j.get('meta') or {}).get('total', 0)
        offset += len(objetos)
        if not objetos or offset >= total:
            break
    logger.info("SciELO (red completa): %d ISSN", len(issns))
    return issns


def _norm(s):
    """Título normalizado: minúsculas, sin acentos ni signos."""
    s = (s or '').lower()
    for orig, rep in zip("áéíóúàãâêôõçñüïî", "aeiouaaaeoocnuii"):
        s = s.replace(orig, rep)
    return re.sub(r'[^a-z0-9 ]+', ' ', s).strip()


def _tokens(s):
    """Palabras significativas del título (sin artículos ni genéricos)."""
    vacias = {'revista', 'de', 'del', 'la', 'el', 'los', 'las', 'y', 'e', 'en',
              'para', 'journal', 'revue', 'da', 'do', 'dos', 'das', 'sobre'}
    return {t for t in _norm(s).split() if t and t not in vacias}


# Palabras demasiado comunes en títulos académicos como para identificar sola
# a una revista: "Revista de Ciencias Sociales" existe en media docena de países.
GENERICAS = {'ciencias', 'sociales', 'social', 'educacion', 'educacao', 'humanidades',
             'investigacion', 'investigacao', 'estudios', 'estudos', 'pedagogia',
             'sociologia', 'politica', 'politicas', 'cultura', 'universidad',
             'universidade', 'ciencia', 'revista', 'anuario', 'cuadernos',
             'internacional', 'latinoamericana', 'iberoamericana', 'nacional'}


def indice_scopus_por_titulo():
    """{conjunto_de_palabras: (issn, eissn, estado)} de Scopus, todas las hojas."""
    import openpyxl
    from indizacion import descargar_scopus
    wb = openpyxl.load_workbook(descargar_scopus(), read_only=True)

    indice = {}
    for hoja in wb.sheetnames:
        h = hoja.lower()
        if not (h.startswith('scopus sources') or h.startswith('discontinued')):
            continue
        discontinuada = h.startswith('discontinued')
        filas = wb[hoja].iter_rows(values_only=True)
        try:
            cab = [str(c or '').strip().lower() for c in next(filas)]
        except StopIteration:
            continue
        if 'source title' not in cab or 'issn' not in cab:
            continue
        i_t, i_i = cab.index('source title'), cab.index('issn')
        i_e = cab.index('eissn') if 'eissn' in cab else None
        i_s = next((i for i, c in enumerate(cab)
                    if c.startswith('active or inactive')), None)

        for f in filas:
            clave = frozenset(_tokens(f[i_t]))
            if not clave:
                continue
            estado = ('Discontinued' if discontinuada
                      else (str(f[i_s] or '').strip() if i_s is not None else 'desconocido'))
            entrada = (f[i_i], f[i_e] if i_e is not None else '', estado, f[i_t])
            if clave not in indice or estado == 'Active':
                indice[clave] = entrada
    wb.close()
    logger.info("Scopus: %d títulos indexados (incluye discontinuados)", len(indice))
    return indice


def buscar_en_scopus(indice, nombre):
    """
    Match por igualdad del conjunto de palabras significativas del título.

    Se probó aceptar que el título de Scopus estuviera *contenido* en el
    nuestro, para capturar casos como "Iconos" ⊂ "Íconos. Revista de Ciencias
    Sociales". Produjo falsos positivos graves: "Cátedra" (Panamá) matcheaba
    con cualquier título de Scopus que contuviera esa palabra, y le asignaba
    un ISSN de otro país. Afirmar una indización falsa es peor que no
    afirmarla, así que la coincidencia debe ser exacta a nivel de palabras.
    Lo que se pierde por título se recupera por ISSN, que es inequívoco.
    """
    clave = frozenset(_tokens(nombre))
    if clave and clave in indice:
        return indice[clave], 1.0

    # Sin el paréntesis desambiguador que agregamos nosotros: "Educación (PUCP)".
    limpio = frozenset(_tokens(re.sub(r'\s*\([^)]*\)\s*$', '', nombre)))
    if limpio and limpio != clave and limpio in indice:
        # Solo si aporta alguna palabra distintiva: "Revista de Ciencias
        # Sociales" a secas describe media docena de revistas distintas.
        if limpio - GENERICAS:
            return indice[limpio], 0.95

    return None, 0.0


def buscar_en_doaj(sesion, nombre):
    """Busca el título en DOAJ. Devuelve ISSN, sitio y confirmación de DOAJ."""
    consulta = requests.utils.quote(_norm(nombre))
    try:
        r = sesion.get(DOAJ_BUSCAR.format(consulta), timeout=30,
                       params={'pageSize': 5})
        if r.status_code != 200:
            return {}
        resultados = r.json().get('results', [])
    except (requests.RequestException, ValueError):
        return {}

    objetivo = _tokens(nombre)
    mejor, punt = None, 0.0
    for it in resultados:
        b = it.get('bibjson', {})
        s = (len(objetivo & _tokens(b.get('title', ''))) /
             max(len(objetivo), len(_tokens(b.get('title', '')) or {1})))
        if s > punt:
            mejor, punt = b, s
    if not mejor or punt < 0.6:
        return {}

    sitio = ''
    for enlace in mejor.get('link', []):
        if enlace.get('type') == 'homepage':
            sitio = enlace.get('url', '')
            break
    return dict(issn_impreso=mejor.get('pissn') or '',
                issn_online=mejor.get('eissn') or '',
                sitio_url=sitio, en_doaj=1, titulo_doaj=mejor.get('title'),
                confianza=punt)


def importar(workers=4, progreso=None):
    init_db()
    indice_scopus = indice_scopus_por_titulo()
    scopus_issn = cargar_scopus()
    scielo = cargar_scielo_red()

    sesion = requests.Session()
    sesion.headers.update(HEADERS)

    resumen = dict(total=len(REVISTAS), nuevas=0, con_issn=0, con_sitio=0,
                   nivel1=0, sin_nivel=0, scopus=0, scielo=0, doaj=0,
                   sin_issn_nombres=[], discrepancias=[])

    def procesar(entrada):
        nombre, pais, institucion, sitio, declarada = entrada
        en_doaj_datos = buscar_en_doaj(sesion, nombre)
        time.sleep(0.08)
        return entrada, en_doaj_datos

    with ThreadPoolExecutor(max_workers=workers) as pool:
        futuros = [pool.submit(procesar, e) for e in REVISTAS]
        for i, fut in enumerate(as_completed(futuros), 1):
            (nombre, pais, institucion, sitio, declarada), doaj = fut.result()

            # Scopus por título (la lista trae ISSN, así que también los aporta).
            sc, punt_sc = buscar_en_scopus(indice_scopus, nombre)
            en_scopus = 1 if sc else 0
            estado_scopus = sc[2] if sc else None

            # ISSN: manda Scopus cuando hubo coincidencia exacta de título;
            # DOAJ solo cuando Scopus no encontró nada.
            #
            # Al revés se producían asignaciones cruzadas: "Práxis Educativa"
            # (UEPG, Brasil) recibía el ISSN de "Praxis educativa (Santa Rosa)"
            # (UNLPam, Argentina), porque la búsqueda de DOAJ por título ignora
            # el acento y devolvía la argentina. Con un ISSN ajeno, todas las
            # indizaciones que se derivan de él quedan mal.
            def _con_guion(v):
                n = re.sub(r'[^0-9Xx]', '', str(v or '')).upper()
                return f"{n[:4]}-{n[4:]}" if len(n) == 8 else ''

            if sc:
                issn_i, issn_o = _con_guion(sc[0]), _con_guion(sc[1])
                if not (issn_i or issn_o):
                    issn_i = doaj.get('issn_impreso') or ''
                    issn_o = doaj.get('issn_online') or ''
            else:
                issn_i = doaj.get('issn_impreso') or ''
                issn_o = doaj.get('issn_online') or ''

            ns = [n for n in (normalizar_issn(issn_i), normalizar_issn(issn_o)) if n]
            if not en_scopus and any(n in scopus_issn for n in ns):
                en_scopus = 1
                estado_scopus = next(scopus_issn[n] for n in ns if n in scopus_issn)

            en_scielo = 1 if any(n in scielo for n in ns) else 0
            sitio_final = sitio or doaj.get('sitio_url') or ''
            if not en_scielo and 'scielo' in sitio_final.lower():
                en_scielo = 1

            nivel = 1 if (en_scopus or en_scielo) else None

            procedencia = []
            if sc:
                procedencia.append(f"Scopus, título exacto: «{sc[3]}»")
            if doaj:
                procedencia.append(f"DOAJ: {doaj.get('titulo_doaj')}")
            if not procedencia:
                procedencia.append("sin coincidencia en Scopus ni DOAJ")

            # Si el ISSN ya está en el catálogo, es la misma revista bajo otro
            # nombre: se enriquece la existente en vez de duplicarla.
            from database import buscar_por_issn, conectar as _conectar
            rid_existente = buscar_por_issn([issn_i, issn_o])
            if rid_existente:
                cx = _conectar()
                cx.execute(
                    """UPDATE revistas
                       SET sitio_url = CASE WHEN COALESCE(sitio_url,'')=''
                                            THEN ? ELSE sitio_url END,
                           indexacion_declarada = COALESCE(indexacion_declarada, ?),
                           pais = COALESCE(pais, ?)
                       WHERE id = ?""",
                    (sitio_final, declarada, pais, rid_existente))
                cx.commit()
                cx.close()
                rid = rid_existente
                estado = 'fusionada'
                resumen.setdefault('fusionadas', 0)
                resumen['fusionadas'] += 1
            else:
                estado, rid = guardar_revista_externa(
                    nombre, pais, institucion, sitio_final, issn_i, issn_o,
                    declarada, " · ".join(procedencia))
                if estado == 'nueva':
                    resumen['nuevas'] += 1

            guardar_indizacion(rid, dict(
                en_scopus=en_scopus, scopus_estado=estado_scopus,
                en_scielo=en_scielo, en_doaj=doaj.get('en_doaj', 0),
                openalex_core=None, h_index=None, works_count=None,
                nivel_conicet=nivel))

            resumen['scopus'] += en_scopus
            resumen['scielo'] += en_scielo
            resumen['doaj'] += doaj.get('en_doaj', 0)
            resumen['con_sitio'] += bool(sitio_final)
            if ns:
                resumen['con_issn'] += 1
            else:
                resumen['sin_issn_nombres'].append(nombre)
            if nivel == 1:
                resumen['nivel1'] += 1
            else:
                resumen['sin_nivel'] += 1

            # Contraste entre lo declarado y lo verificado.
            if 'scopus' in (declarada or '').lower() and not en_scopus:
                resumen['discrepancias'].append(
                    f"{nombre} ({pais}): declara Scopus" +
                    (f", ISSN {issn_i or issn_o} no está en la lista de Elsevier"
                     if ns else ", y no se pudo obtener su ISSN para verificar"))

            if progreso and i % 10 == 0:
                progreso(i, len(REVISTAS))

    registrar_actualizacion(
        "externas",
        f"{resumen['total']} revistas externas: {resumen['nivel1']} Nivel 1, "
        f"{resumen['scopus']} Scopus, {resumen['scielo']} SciELO, "
        f"{resumen['doaj']} DOAJ, {len(resumen['discrepancias'])} sin confirmar Scopus")
    return resumen


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s %(message)s')
    r = importar(progreso=lambda i, t: print(f"  {i}/{t}", flush=True))
    print("\nRESUMEN")
    for k in ('total', 'nuevas', 'con_issn', 'con_sitio', 'nivel1', 'sin_nivel',
              'scopus', 'scielo', 'doaj'):
        print(f"  {k}: {r[k]}")
    if r['sin_issn_nombres']:
        print(f"\n  SIN ISSN ({len(r['sin_issn_nombres'])}) — no verificables:")
        for n in sorted(r['sin_issn_nombres']):
            print("    -", n)
    if r['discrepancias']:
        print(f"\n  DECLARAN SCOPUS SIN CONFIRMAR ({len(r['discrepancias'])}):")
        for d in sorted(r['discrepancias']):
            print("    -", d)
